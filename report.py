import time
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
from aiogram.types import InputMediaDocument, BufferedInputFile
from sqlalchemy import select, and_, func

from db import AsyncSessionLocal
from main import bot
from message_processing import to_date
from models.application import Application
from models.payment import Payment
from models.subscription import Subscription
from models.user import User
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


async def generate_report():
    async with AsyncSessionLocal() as session:
        async with session.begin():
            result = await session.execute(
                select(User).filter(User.admin == True)
            )

            users = result.scalars().all()

            now = datetime.now()
            yesterday_start = datetime(now.year, now.month, now.day) - timedelta(days=1)
            yesterday_end = datetime(now.year, now.month, now.day) + timedelta(days=1)

            yesterday_start_timestamp = int(yesterday_start.timestamp() * 1000)
            yesterday_end_timestamp = int(yesterday_end.timestamp() * 1000)

            report = await collect_data(session, yesterday_start_timestamp, yesterday_end_timestamp)

            await send_report(report, users, "За этот день нет новых заявок")

            if now.day == 1:
                last_month_end = datetime(now.year, now.month, 1) - timedelta(seconds=1)
                last_month_start = datetime(last_month_end.year, last_month_end.month, 1)

                last_month_end_timestamp = int(last_month_end.timestamp() * 1000)
                last_month_start_timestamp = int(last_month_start.timestamp() * 1000)

                report = await collect_data(session, last_month_start_timestamp, last_month_end_timestamp)

                await send_report(report, users, "За этот месяц нет новых заявок")


async def send_report(report, users, text_none):
    if report:
        for user in users:
            try:
                await bot.send_media_group(
                    chat_id=user.telegram_chat_id,
                    media=[report]
                )
            except:
                ...
    else:
        for user in users:
            try:
                await bot.send_message(
                    chat_id=user.telegram_chat_id,
                    text=text_none
                )
            except:
                ...


async def collect_data(session, start_unix, end_unix):

    df_app = await collect_application_data(session, start_unix, end_unix)

    df_user = await collect_user_data(session, start_unix, end_unix)

    if df_app is None and df_user is None:
        return None

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if df_app is not None:
            df_app.to_excel(writer, index=False, sheet_name='Заявки')
        if df_user is not None:
            df_user.to_excel(writer, index=False, sheet_name='Пользователи')

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]

            for col in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = 30

            first_col_letter = get_column_letter(1)
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet[f'{first_col_letter}{row}']
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet[f'{get_column_letter(col)}1']
                header_cell.value = header_cell.value.replace('\\n', '\n')
                header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    output.seek(0)
    excel_bytes = output.getvalue()

    if end_unix - start_unix >= 86400000 * 1.3:
        date_str = datetime.fromtimestamp(end_unix / 1000 - 60).strftime("%m-%Y")
    else:
        date_str = datetime.fromtimestamp(end_unix / 1000 - 60).strftime("%d-%m-%Y")

    report = InputMediaDocument(
        media=BufferedInputFile(
            excel_bytes,
            filename=f'report_{date_str}.xlsx'
        ),
    )

    return report


async def collect_user_data(session, start_unix, end_unix):
    result = await session.execute(
        select(User).filter(
            and_(
                User.created >= start_unix,
                User.created <= end_unix,
            )
        )
    )
    new_users = result.scalars().all()

    if len(new_users) == 0:
        return None

    num, create_user, fio, phone, pay_date, next_pay_date, is_payed, pay_price = [], [], [], [], [], [], [], []

    user_ids = [int(user.telegram_user_id) for user in new_users]

    result = await session.execute(
        select(Subscription).filter(Subscription.telegram_user_id.in_(user_ids))
    )
    subscriptions = result.scalars().all()

    subquery = (
        select(
            Payment.telegram_user_id,
            func.max(Payment.created).label('created')
        )
        .filter(Payment.telegram_user_id.in_(user_ids))
        .group_by(Payment.telegram_user_id)
        .subquery()
    )

    stmt = (
        select(Payment)
        .join(subquery,
              and_(
                  Payment.telegram_user_id == subquery.c.telegram_user_id,
                  Payment.created == subquery.c.created
              ))
    )

    result = await session.execute(stmt)
    last_payments = result.scalars().all()

    for u in new_users:
        num.append(u.id)
        create_user.append(to_date(u.created))
        fio.append(u.name)
        phone.append(u.phone)

        if u.admin:
            pay_date.append("НЕТ")
            is_payed.append("НЕТ")
            next_pay_date.append("-")
            pay_price.append(0)
        else:
            subscription = next((obj for obj in subscriptions if obj.telegram_user_id == u.telegram_user_id), None)
            last_payment = next((obj for obj in last_payments if obj.telegram_user_id == u.telegram_user_id), None)

            if last_payment:
                pay_date.append(to_date(last_payment.created, only_date=True))
            else:
                pay_date.append("-")

            if subscription.end_time > int(time.time() * 1000):

                is_payed.append("ДА")

                next_pay_date.append(to_date(subscription.end_time, only_date=True))

                if subscription.status == 3:
                    pay_price.append(0)
                elif subscription.status == 0:
                    pay_price.append(last_payment.amount if last_payment else 0)
                else:
                    pay_price.append("-")
            else:
                is_payed.append("НЕТ")
                next_pay_date.append("-")
                pay_price.append("-")

    data = {
        "№": num,
        "Дата регистрации\nИсполнителя": create_user,
        "Ф.И.О.\nИсполнителя": fio,
        "Телефон\nИсполнителя": phone,
        "Даты оплаты подписки\nИсполнителем": pay_date,
        "Дата следующей оплата подпистки\nИсполнителем": next_pay_date,
        "Оплата подписки\nДА / НЕТ": is_payed,
        "Оплата подписки, руб.": pay_price
    }

    df = pd.DataFrame(data)

    return df


async def collect_application_data(session, start_unix, end_unix):
    result = await session.execute(
        select(Application).filter(
            and_(
                Application.created >= start_unix,
                Application.created <= end_unix,
            )
        )
    )
    applications = result.scalars().all()

    if len(applications) == 0:
        return None

    (num, item_name, city, reg_num, fio, phone, customer_name, first_message, last_message, create_date,
     last_message_date, close_date) = [], [], [], [], [], [], [], [], [], [], [], []

    working_ids = [application.working_user_id for application in applications]

    result = await session.execute(
        select(User).filter(User.telegram_user_id.in_(working_ids))
    )
    users = result.scalars().all()

    for i, application in enumerate(applications):
        user = next((u for u in users if u.telegram_user_id == application.working_user_id), None)

        num.append(application.id)

        if user is None:
            reg_num.append("-")
            phone.append("-")
            fio.append("-")
        else:
            reg_num.append(user.id)
            phone.append(user.phone)
            fio.append(user.name)

        item_name.append(application.item_name)
        city.append(application.item_location)
        customer_name.append(application.username)
        first_message.append(application.content)
        last_message.append(application.last_message_text)
        create_date.append(to_date(application.created))
        last_message_date.append(to_date(application.last_message_time))
        close_date.append(to_date(application.close_app_time) if application.close_app_time != -1 else "-")

    data = {
        "№": num,
        "Наименование объявления": item_name,
        "Город": city,
        "Номер регистрации\nИсполнителя": reg_num,
        "Ф.И.О.\nИсполнителя": fio,
        "Телефон\nИсполнителя": phone,
        "Наименование\nЗаказчика": customer_name,
        "Первое сообщение\nЗаказчика": first_message,
        "Последнее сообщение\nЗаказчика": last_message,
        "Дата обращения\nЗаказчика": create_date,
        "Дата последнего сообщения": last_message_date,
        "Дата закрытия заявки": close_date
    }

    df = pd.DataFrame(data)

    return df
